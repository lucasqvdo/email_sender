import win32com.client as win32
import openpyxl 
import pyautogui as pg
import webbrowser as web
import time

def autorizaçoes(emaildestino, codigo_loja, data, serviço, marca, shopping, horario, empresa, nome_colaboradora, rg_colaboradora, cpf_colaboradora):
    
    # criar a integração com o outlook
    outlook = win32.Dispatch('outlook.application')

    # criar um email
    email = outlook.CreateItem(0)


    # configurar as informações do seu e-mail

    email.To = emaildestino
    email.Subject = "Autorização de acesso "+ codigo_loja+" / "+ str(data) + " / " + serviço
    email.HTMLBody = f"""
    <p>Olá,  Equipe <b> {marca} {shopping}!</b></p>

    <p>Por gentileza, providenciar a autorização de acesso junto à administração do Shopping, seguindo as orientações abaixo:</p>


    <p style="color: red;"><b>Aguardo a confirmação de acesso via e-mail.</b></p>

    <p><b>Data:</b>{data} <br>
    <b>Horário:</b>{horario} <br>
    <b>Serviço:</b> {serviço}</p>


    <p><b>Dados dos prestadores de serviço:</b></p>
    <b>Empresa:</b> {empresa}<br>   
    <b>Nome:</b> {nome_colaboradora}<br>
    <b>RG:</b> {rg_colaboradora} <br> 
    <b>CPF:</b> {cpf_colaboradora}     <br>  

    <p>Observações:</p>
    <p style="color:blue"><b>Será necessário que uma funcionária da loja receba a prestadora no horário estabelecido para iniciar a limpeza.
    <b/></p>
    <p>Orientem a prestadora, identificando os locais onde necessitam de maior cuidado e atenção, tais como:<br>
    - Limpeza de piso, rodapés, parede e molduras de parede. <br>
    - Limpeza de móveis. <br>
    - Limpeza de luminárias (necessário utilizar escada alta, antes das 10h).<br>
    - Limpeza pontual de vidros e espelhos.<br>
    - Limpeza do Estoque.<br>
    - A tinta das paredes é lavável, portanto, pode ser usada uma esponja macia com detergente neutro para limpeza.<br>
    - O piso da loja é de porcelanato e dos provadores é laminado.<br>

    <br>
    <b>Obrigada!</b>
    </p>
    """

    # anexo = "C://Users/joaop/Downloads/arquivo.xlsx"
    # email.Attachments.Add(anexo)

    email.Send()
    print("Solicitação de autorização enviada via email para a loja "+ codigo_loja)
        
def invite(serviço, codigo_loja, marca, shopping, emaildestino, datahora):
    
    outlook = win32.Dispatch('outlook.application')
    appt = outlook.CreateItem(1) # AppointmentItem
    appt.Start = str(datahora) # yyyy-MM-dd hh:mm
    appt.Subject = serviço + " / " + codigo_loja
    appt.Duration = 480 # In minutes (60 Minutes)
    appt.Location = codigo_loja+" / "+ marca +" / "+ shopping
    appt.MeetingStatus = 1 # 1 - olMeeting; Changing the appointment to meeting. Only after changing the meeting status recipients can be added
    #appt.Organizer = sender
    appt.ReminderMinutesBeforeStart = 15
    appt.ResponseRequested = True
    appt.Recipients.Add(emaildestino) # Don't end ; as delimiter

    # Set Pattern, to recur every day, for the next 5 days
    #pattern = appt.GetRecurrencePattern()
    #pattern.RecurrenceType = 0
    #pattern.Occurrences = "5"

    appt.Save()
    appt.Send()

    print('Convite para reunião enviado referente a ' + serviço +'loja' +codigo_loja)

def whatsapp(lista_nomes): 
    

    #web.open("https://web.whatsapp.com")
    first= True
    time.sleep(4)
    

    for nome in lista_nomes:
    
            
        print("pagina aberta")   
        if first:
            time.sleep(6)
            first=False
        width, height = pg.size()
        pg.click(width/2, height/2)
        time.sleep(6)
        pg.press('tab')
        pg.press('tab')
        pg.press('tab')
        pg.press('tab')
        pg.press('tab')
        pg.press('tab')
        pg.press('tab')
        pg.press('tab')
        print("nome a ser digitado" + nome)
        pg.typewrite(nome, interval=0.3)
        time.sleep(2) 
        pg.press('enter')

        time.sleep(2) 
        pg.press('enter')

        time.sleep(1)
        pg.typewrite('Essa é uma mensagem automática', interval=0.30)
        pg.press('enter')
        time.sleep(1)
        pg.press('esc')
        
        print("Mensagem enviada para "+nome)
        time.sleep(5)



    else:
        pg.hotkey('ctrl','w')
      

while True:


    workbook_cronograma = openpyxl.load_workbook('AUTORIZAÇÕES DE ACESSO.xlsx')
    sheet_cronograma = workbook_cronograma['Plan1']


    print("Seja bem vindo ao criador de autorizações e invites!")


    resposta=input("O que você deseja fazer? \n 1 - Enviar solicitações de autorização \n 2 - Enviar invites \n 3 - Cobrar autorizações via WhatsApp \n \n")

    if int(resposta) == 1:
        contador = 0



        for linha in sheet_cronograma.iter_rows(min_row=2):
            enviado=linha[11].value
            

            if enviado == "não":
                marca=linha[0].value
                codigo_loja=linha[1].value
                shopping=linha[2].value
                data=linha[3].value
                horario=linha[4].value
                serviço=linha[5].value
                emaildestino=linha[6].value
                empresa=linha[7].value
                nome_colaboradora=linha[8].value
                rg_colaboradora=linha[9].value
                cpf_colaboradora=linha[10].value

                
                
                autorizaçoes(emaildestino, codigo_loja, data, serviço, marca, shopping, horario, empresa, nome_colaboradora, rg_colaboradora, cpf_colaboradora)
                linha[11].value = 'sim'
                workbook_cronograma.save(filename="AUTORIZAÇÕES DE ACESSO.xlsx") 
                workbook_cronograma.close()
                print("Planilha atualizada")
                contador += 1
            
        if contador == 0:
            print("\n \nNenhuma solicitação foi enviada. \n \n")

        else:
            print("\n \nForam enviadas "+ str(contador) +" solicitações. \n \n")     
          
    elif int(resposta) == 2:
        contador = 0
        for linha in sheet_cronograma.iter_rows(min_row=2):
            enviado=linha[12].value
            
            if enviado == "não":
                marca=linha[0].value
                codigo_loja=linha[1].value
                shopping=linha[2].value
                data=linha[3].value
                horario=linha[4].value
                serviço=linha[5].value
                emaildestino=linha[6].value
                empresa=linha[7].value
                nome_colaboradora=linha[8].value
                rg_colaboradora=linha[9].value
                cpf_colaboradora=linha[10].value
                datahora = str(data)[0:10]+" "+str(horario)

               
                invite(serviço, codigo_loja, marca, shopping, emaildestino, datahora)
                linha[12].value = 'sim'
                workbook_cronograma.save(filename="AUTORIZAÇÕES DE ACESSO.xlsx") 
                workbook_cronograma.close()
                contador += 1
            
        if contador == 0:
            print("\n \nNenhum invite foi enviado. \n \n")
            

        else:
            print("\n \nForam enviados "+ str(contador) +" invites. \n \n")

    elif int(resposta)==3:
        contador=0
        lista_nomes=[]


        for linha in sheet_cronograma.iter_rows(min_row=2):
            enviado=linha[13].value
            
            if enviado == "não":
                codigo_loja=linha[1].value
                
                lista_nomes.append(codigo_loja)
                linha[13].value = 'sim'
                workbook_cronograma.save(filename="AUTORIZAÇÕES DE ACESSO.xlsx") 
                workbook_cronograma.close()
                contador +=1
        web.open("https://web.whatsapp.com/send/?phone="+str(5511984489030)+"&text="+"mensagem de teste")
        time.sleep(5)
        pg.press('enter')
        #whatsapp(lista_nomes)
        if contador == 0:
            print("\n \nNenhuma mensagem foi enviada. \n \n")

        elif contador == 1:

            print("Uma mensagem foi enviada.")            

        else:
            print("\n \nForam enviadas "+ str(contador) +" mensagens. \n \n")
